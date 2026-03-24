#!/usr/bin/env python3
"""
compare_excel.py — Compara dos archivos Excel generados por pdf_to_excel.py.

Verifica:
  - Cantidad de filas de movimientos
  - Suma de DÉBITOS
  - Suma de CRÉDITOS
  - Detalle de filas con diferencias (con --detalle)

Uso:
    python3 compare_excel.py original.xlsx nuevo.xlsx [--detalle]
"""

import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: pip install openpyxl")
    sys.exit(1)

SHEET        = "Movimientos"
COL_FECHA    = 1
COL_CONCEPTO = 2
COL_DEBITOS  = 7
COL_CREDITOS = 8


def _find_header_row(ws):
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str) and cell.value.strip().lower() == "fecha":
            return cell.row
    return None


def extract_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"Hoja '{SHEET}' no encontrada en {path}")
    ws = wb[SHEET]

    hdr = _find_header_row(ws)
    if hdr is None:
        raise ValueError(f"No se encontró fila de encabezado en {path}")

    rows = []
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        if row[COL_FECHA - 1] is None:
            continue
        rows.append({
            "fecha":    row[COL_FECHA    - 1],
            "concepto": row[COL_CONCEPTO - 1],
            "debitos":  row[COL_DEBITOS  - 1],
            "creditos": row[COL_CREDITOS - 1],
        })

    wb.close()
    return rows


def extract_stats(rows):
    return {
        "filas":    len(rows),
        "debitos":  sum(abs(r["debitos"])  for r in rows if isinstance(r["debitos"],  (int, float))),
        "creditos": sum(abs(r["creditos"]) for r in rows if isinstance(r["creditos"], (int, float))),
    }


def compare(path_a, path_b, detalle=False):
    rows_a = extract_rows(path_a)
    rows_b = extract_rows(path_b)
    a = extract_stats(rows_a)
    b = extract_stats(rows_b)

    print(f"\n{'Métrica':<20} {'Original':>18} {'Nuevo':>18} {'Diff':>12}")
    print("-" * 72)

    ok = True
    for key, label, fmt in [
        ("filas",    "Filas",         lambda v: str(v)),
        ("debitos",  "Suma DÉBITOS",  lambda v: f"{v:,.2f}"),
        ("creditos", "Suma CRÉDITOS", lambda v: f"{v:,.2f}"),
    ]:
        va, vb = a[key], b[key]
        diff = vb - va
        match = "✓" if abs(diff) < 0.01 else "✗ DIFERENCIA"
        if abs(diff) >= 0.01:
            ok = False
        print(f"{label:<20} {fmt(va):>18} {fmt(vb):>18} {fmt(diff):>10}  {match}")

    print()
    if ok:
        print("✓ Los archivos son equivalentes.")
    else:
        print("✗ Hay diferencias. Revisar el output.")

    if detalle and not ok:
        _mostrar_diferencias(rows_a, rows_b)

    return ok


def _mostrar_diferencias(rows_a, rows_b):
    print(f"\n{'─'*80}")
    print("DETALLE DE FILAS CON DIFERENCIAS EN DÉBITOS")
    print(f"{'─'*80}")
    print(f"{'#':<5} {'Fecha':<8} {'Concepto':<32} {'Original':>14} {'Nuevo':>14} {'Diff':>12}")
    print(f"{'─'*80}")

    total_diff = 0.0
    for i, (ra, rb) in enumerate(zip(rows_a, rows_b), 1):
        da = ra["debitos"] if isinstance(ra["debitos"], (int, float)) else 0.0
        db = rb["debitos"] if isinstance(rb["debitos"], (int, float)) else 0.0
        diff = db - da
        if abs(diff) >= 0.01:
            concepto = str(rb.get("concepto") or ra.get("concepto") or "")[:32]
            print(
                f"{i:<5} {str(ra['fecha']):<8} {concepto:<32} "
                f"{da:>14,.2f} {db:>14,.2f} {diff:>12,.2f}"
            )
            total_diff += diff

    if len(rows_a) != len(rows_b):
        print(f"\n  Filas extra en nuevo ({len(rows_b) - len(rows_a):+d}) — comparación fila a fila truncada.")

    print(f"{'─'*80}")
    print(f"  Diferencia total en débitos: {total_diff:,.2f}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Compara dos Excel de extractos bancarios")
    parser.add_argument("original")
    parser.add_argument("nuevo")
    parser.add_argument("--detalle", action="store_true",
                        help="Muestra las filas específicas con diferencias")
    args = parser.parse_args()

    for p in (args.original, args.nuevo):
        if not Path(p).exists():
            print(f"Error: no existe '{p}'")
            sys.exit(1)

    ok = compare(args.original, args.nuevo, detalle=args.detalle)
    sys.exit(0 if ok else 1)
