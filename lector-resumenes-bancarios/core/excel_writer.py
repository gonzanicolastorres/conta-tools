"""
core/excel_writer.py — Generación del archivo Excel de salida.
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    raise ImportError(f"Dependencia faltante: {e}. Instalá con: pip install openpyxl")

_DEFAULT_COLUMNS = ["fecha", "concepto", "f_valor", "comprobante", "origen", "canal", "debitos", "creditos", "saldos"]

# Ancho de columna por defecto según el tipo de campo
_WIDE_TEXT  = 50    # columnas de descripción larga
_WIDE_NUM   = 18    # columnas numéricas
_WIDE_CODE  = 12    # códigos cortos
_WIDE_SMALL = 8

def _col_width(col_name):
    name = col_name.lower()
    if "concepto" in name or "descripcion" in name or "detalle" in name:
        return _WIDE_TEXT
    if any(k in name for k in ("debito", "credito", "saldo", "importe", "monto")):
        return _WIDE_NUM
    if any(k in name for k in ("canal", "origen", "tipo")):
        return _WIDE_SMALL
    return _WIDE_CODE

def _col_header(col_name):
    """Convierte clave interna a label legible: 'f_valor' → 'F.Valor'"""
    return col_name.replace("_", ".").title()


def write_excel(metadata, transactions, output_path, columns=None, empresa=None):
    """
    Genera el Excel con dos hojas:
      - Movimientos: tabla formateada con encabezado y colores.
      - OCR Raw: línea cruda por movimiento para diagnóstico.

    columns: lista de nombres de columna en orden (de la calibración).
             Si es None se usa el set por defecto de ICBC.
    """
    if columns is None:
        columns = _DEFAULT_COLUMNS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    COLOR_HEADER    = "1A3A5C"
    COLOR_SUBHEADER = "2E6DA4"
    COLOR_DEBIT     = "FDECEA"
    COLOR_CREDIT    = "E8F5E9"
    COLOR_ALT       = "F5F8FC"
    COLOR_EMPTY     = "FFFF99"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style(cell, bold=False, bg=None, fg="000000", align="left", fmt=None):
        cell.font = Font(bold=bold, color=fg, size=10)
        if bg:
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border
        if fmt:
            cell.number_format = fmt

    # ── Encabezado resumen (se re-mergea más abajo con el ancho correcto) ─────
    banco    = metadata.get("banco", "")
    titulo   = f"{empresa}  —  Extracto de Cuenta Corriente" if empresa else "Extracto de Cuenta Corriente"
    if banco:
        titulo += f"  [{banco.upper()}]"
    ws["A1"] = titulo
    style(ws["A1"], bold=True, bg=COLOR_HEADER, fg="FFFFFF", align="center")
    ws.row_dimensions[1].height = 22

    info = [
        ("Empresa:",      empresa or ""),
        ("Titular:",      metadata.get("titular", "")),
        ("CUIT:",         metadata.get("cuit", "")),
        ("Período:",      metadata.get("periodo", "")),
        ("Saldo inicial:", metadata.get("saldo_inicial", "")),
    ]
    for i, (label, value) in enumerate(info, 2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        c = ws.cell(row=i, column=2, value=value)
        c.font = Font(size=10)
        if label == "Saldo inicial:" and value:
            c.number_format = "#.##0,00"

    # ── Encabezados de columnas ───────────────────────────────────────────────
    HDR_ROW = len(info) + 3
    n_cols = len(columns) + 1  # +1 por la columna Hoja
    last_col_letter = openpyxl.utils.get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")  # re-merge con el ancho correcto

    headers    = [_col_header(c) for c in columns] + ["Hoja"]
    col_widths = [_col_width(c)  for c in columns] + [6]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        style(c, bold=True, bg=COLOR_SUBHEADER, fg="FFFFFF", align="center")
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[HDR_ROW].height = 18

    # ── Datos ─────────────────────────────────────────────────────────────────
    NUM_FMT = "#.##0,00"

    for row_i, tx in enumerate(transactions, HDR_ROW + 1):
        alt = (row_i - HDR_ROW) % 2 == 0
        debit_cols  = [c for c in columns if any(k in c.lower() for k in ("debito",  "cargo",  "egreso"))]
        credit_cols = [c for c in columns if any(k in c.lower() for k in ("credito", "abono",  "ingreso"))]
        has_debit   = any(tx.get(c + "_num") is not None for c in debit_cols)
        has_credit  = any(tx.get(c + "_num") is not None for c in credit_cols)
        if not has_debit and not has_credit:
            bg = COLOR_EMPTY
        elif has_debit:
            bg = COLOR_DEBIT
        elif has_credit:
            bg = COLOR_CREDIT
        elif alt:
            bg = COLOR_ALT
        else:
            bg = None

        values = []
        aligns = []
        fmts   = []
        for col_name in columns:
            is_numeric = any(k in col_name.lower() for k in ("debito", "credito", "saldo", "importe", "monto"))
            key = col_name + "_num" if is_numeric else col_name
            values.append(tx.get(key))
            aligns.append("right" if is_numeric else ("left" if _col_width(col_name) >= _WIDE_TEXT else "center"))
            fmts.append(NUM_FMT if is_numeric else None)
        # columna Hoja
        values.append(tx["pagina"])
        aligns.append("center")
        fmts.append(None)

        for col, (val, align, fmt) in enumerate(zip(values, aligns, fmts), 1):
            c = ws.cell(row=row_i, column=col, value=val)
            style(c, bg=bg, align=align, fmt=fmt)

        ws.row_dimensions[row_i].height = 15

    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # ── Hoja OCR Raw ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("OCR Raw")
    ws2["A1"] = "Pg"
    ws2["B1"] = "Línea raw"
    for i, tx in enumerate(transactions, 2):
        ws2.cell(row=i, column=1, value=tx["pagina"])
        raw = " | ".join(f"{c}={tx[c]}" for c in columns if tx.get(c))
        ws2.cell(row=i, column=2, value=raw)
    ws2.column_dimensions["B"].width = 120

    # ── Hoja Alertas ──────────────────────────────────────────────────────────
    # Detectar filas vacías (sin debitos ni creditos numéricos válidos)
    filas_vacias = [
        (tx["pagina"], tx.get("fecha", ""), tx.get("concepto", ""))
        for tx in transactions
        if not any(
            isinstance(tx.get(c + "_num"), (int, float))
            for c in columns
            if any(k in c.lower() for k in ("debito", "credito"))
        )
    ]

    if filas_vacias:
        ws_alertas = wb.create_sheet("Alertas")
        ws_alertas["A1"] = "Filas sin monto (revisar manualmente)"
        ws_alertas["A1"].font = Font(bold=True, color="FFFFFF")
        ws_alertas["A1"].fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
        ws_alertas.merge_cells("A1:C1")
        for i, (pag, fecha, concepto) in enumerate(filas_vacias, 2):
            ws_alertas.cell(row=i, column=1, value=pag)
            ws_alertas.cell(row=i, column=2, value=fecha)
            ws_alertas.cell(row=i, column=3, value=concepto)
        ws_alertas.column_dimensions["B"].width = 15
        ws_alertas.column_dimensions["C"].width = 60

    wb.save(output_path)

